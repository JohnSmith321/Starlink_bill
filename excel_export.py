import zipfile
from datetime import datetime
from rich.console import Console
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import EXCEL_HEADERS, REPORT_HEADERS, EXCEL_FILE, ZIP_FILE, INVOICES_DIR

console = Console()


def build_excel(records: list[dict], run_label: str = "export"):
    """
    Write all invoice records to a formatted Excel workbook.
    Creates two sheets:
      - Invoices : one row per invoice
      - Summary  : totals by currency
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # ── Styles ──────────────────────────────────────────────────────
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1A73E8")
    data_font    = Font(name="Arial", size=10)
    alt_fill     = PatternFill("solid", start_color="F0F4FF")
    border_side  = Side(style="thin", color="D0D7E3")
    cell_border  = Border(bottom=border_side)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")

    col_widths = [28, 32, 18, 24, 14, 10, 45, 38]

    # ── Header row ───────────────────────────────────────────────────
    for col, (header, width) in enumerate(zip(EXCEL_HEADERS, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 24

    # ── Data rows ────────────────────────────────────────────────────
    for r_idx, rec in enumerate(records, start=2):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill()
        values = [
            rec.get("customer_account", ""),
            rec.get("invoice_number",   ""),
            rec.get("invoice_date",     ""),
            rec.get("payment_date",     ""),
            rec.get("amount",           0.0),
            rec.get("currency",         ""),
            rec.get("product",          ""),
            rec.get("pdf_file",         ""),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col, value=value)
            cell.font      = data_font
            cell.border    = cell_border
            cell.fill      = fill
            cell.alignment = center_align if col in (3, 4, 5, 6) else left_align

    # ── Summary sheet ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14)

    ws2["A3"] = "Total Invoices"
    ws2["B3"] = "=COUNTA(Invoices!B2:B10000)"

    ws2["A4"] = "Total Amount (PHP)"
    ws2["B4"] = '=SUMIF(Invoices!F2:F10000,"PHP",Invoices!E2:E10000)'

    ws2["A5"] = "Total Amount (USD)"
    ws2["B5"] = '=SUMIF(Invoices!F2:F10000,"USD",Invoices!E2:E10000)'

    ws2["A7"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws2["A7"].font = Font(name="Arial", italic=True, color="888888")

    for col in ["A", "B"]:
        ws2.column_dimensions[col].width = 28

    out_path = EXCEL_FILE.parent / f"starlink_invoices_{run_label}.xlsx"
    wb.save(str(out_path))
    console.print(f"[green]✓ Excel saved → {out_path}[/green]")
    return out_path


def zip_pdfs(run_label: str = "export", pdf_dir=None):
    """Zip all PDFs in pdf_dir into a run-specific archive next to the Excel file."""
    from pathlib import Path
    if pdf_dir is None:
        pdf_dir = INVOICES_DIR
    pdf_dir = Path(pdf_dir)
    pdf_files = list(pdf_dir.rglob("*.pdf"))
    if not pdf_files:
        console.print("[yellow]No PDFs to zip.[/yellow]")
        return None
    out_path = ZIP_FILE.parent / f"starlink_invoices_{run_label}.zip"
    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for pdf in pdf_files:
            # Store with relative path inside zip: mode_slug/filename.pdf
            arcname = pdf.relative_to(pdf_dir.parent)
            zf.write(pdf, arcname)
    console.print(f"[green]✓ Zipped {len(pdf_files)} PDFs → {out_path}[/green]")
    return out_path


def build_report_excel(rows: list[dict], run_label: str = "report"):
    """
    Write account status report to a formatted Excel workbook.
    One row per account with billing + subscription info.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Account Report"

    # ── Styles ──────────────────────────────────────────────────────
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="0D47A1")
    data_font    = Font(name="Arial", size=10)
    alt_fill     = PatternFill("solid", start_color="E3F2FD")
    border_side  = Side(style="thin", color="B0BEC5")
    cell_border  = Border(bottom=border_side)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    col_widths = [24, 22, 16, 28, 60, 14, 12, 16, 20, 16, 20, 16]

    # ── Header row ───────────────────────────────────────────────────
    for col, (header, width) in enumerate(zip(REPORT_HEADERS, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28

    # ── Data rows ────────────────────────────────────────────────────
    field_keys = [
        "account", "account_id", "balance_due", "billing_cycle",
        "subscription_status", "service_plan_status", "ocean_mode",
        "device_status", "serial_no", "uptime", "software_version", "wifi_status",
    ]
    for r_idx, rec in enumerate(rows, start=2):
        fill = alt_fill if r_idx % 2 == 0 else PatternFill()
        for col, key in enumerate(field_keys, start=1):
            cell = ws.cell(row=r_idx, column=col, value=rec.get(key, ""))
            cell.font      = data_font
            cell.border    = cell_border
            cell.fill      = fill
            cell.alignment = left_align if col in (5,) else center_align

    # ── Footer ───────────────────────────────────────────────────────
    footer_row = len(rows) + 3
    ws.cell(row=footer_row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=footer_row, column=1).font = Font(name="Arial", italic=True, color="888888")

    out_path = EXCEL_FILE.parent / f"starlink_report_{run_label}.xlsx"
    wb.save(str(out_path))
    console.print(f"[green]✓ Report saved → {out_path}[/green]")
    return out_path
